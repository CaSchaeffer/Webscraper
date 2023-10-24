import requests
from bs4 import BeautifulSoup
import openpyxl

# Define the base URL of the website
base_url = "https://avagrar.de/search?sSearch="

# Open the Excel file
workbook = openpyxl.load_workbook("C:\Python\RLHCrawler\item_data.xlsx")

# Assuming the search terms are in the second column of the first sheet 
sheet = workbook.active
search_terms = [cell.value for cell in sheet['B'] if cell.value is not None]

# Loop through the search terms, searches for found value in the excel sheet
for row_index, search_term in enumerate(search_terms, start=1):
    # Create the complete URL with the search term
    url = f"{base_url}{search_term}"

    # Send an HTTP GET request to the URL
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content of the page
        soup = BeautifulSoup(response.text, "html.parser")

        # Find and extract the title of the product, productname
        product_title_element = soup.find("a", class_="product--title")
        if product_title_element:
            product_title = product_title_element["title"]
        else:
            product_title = "Product title not found on the page."

        additional_info_element = soup.find("div", class_="price--unit")
        if additional_info_element:
            additional_text = additional_info_element["title"]
        else:
            additional_text = "Unit information not found on the page."

        # Find the element containing the unit of mesurement
        searched_item_value = soup.find("span", class_="price--default is--nowrap")

        if searched_item_value:
            # Extract the text from the element to get the price
            value = searched_item_value.get_text(strip=True)
        else:
            value = f"Searched item value for '{search_term}' not found on the page."

        # Insert the found data in the same row as the search term but in the next colums
        sheet.cell(row=row_index, column=3, value=product_title)
        sheet.cell(row=row_index, column=4, value=additional_text)
        sheet.cell(row=row_index, column=5, value=value)

    else:
        print(f"Failed to retrieve the web page for search term '{search_term}'. Status code:", response.status_code)
# Save the modified Excel file
workbook.save("C:\Python\RLHCrawler\item_data_updated.xlsx")
print("Done Scraping and saved the Data, Boss")